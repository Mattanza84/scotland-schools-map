"""
Builds data/schools.json from the raw source files in data/raw/.

Sources (see data/SOURCES.md for full provenance):
  - data/raw/schools_scotland_page{1,2}.geojson
      School locations, sector, denomination, address. Fetched from the
      "Schools_Scotland__2022" ArcGIS FeatureServer layer (WGS84).
  - data/raw/foi_inspections_2025.xlsx
      Education Scotland inspection quality-indicator (QI) grades, 1 (weakest)
      to 6 (strongest), covering all inspections up to 1 April 2025.
      Response to FOI-202500457731. Used as the rating for primary schools,
      and as a fallback for secondary schools with no attainment data.
  - data/raw/school_level_stats_2025.xlsx
      Scottish Government school-level summary statistics 2025, providing
      up-to-date pupil rolls and teacher FTE per school.
  - data/raw/breadth_and_depth_2023-24.json, breadth_and_depth_2024-25.json
      SQA attainment (% of leavers with 5+ awards at Higher level or above),
      from the Scottish Government "Schools - Breadth and Depth of
      Qualifications" statistics.gov.scot dataset. Used as the rating for
      secondary schools where available -- a more current, direct measure
      than the sparse/dated inspection data, and secondary schools have no
      equivalent to primary's ACEL attainment stats at school level.

Run with: python3 scripts/build_schools_json.py
"""
import json
import os
import re
from datetime import date

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(ROOT, "data", "raw")
OUT_PATH = os.path.join(ROOT, "data", "schools.json")

SECTOR_MAP = {
    "Primary": "primary",
    "Secondary": "secondary",
}

# Education Scotland's own wording for the 1-6 scale, used to label the
# averaged score. This is an illustrative average across whichever quality
# indicators a given inspection graded, not an official single overall grade
# (Education Scotland does not publish one).
GRADE_LABELS = {
    6: "Excellent",
    5: "Very Good",
    4: "Good",
    3: "Satisfactory",
    2: "Weak",
    1: "Unsatisfactory",
}

# Bands for secondary schools' SQA attainment percentage (% of leavers with
# 5+ awards at Higher level or above), reusing the same six labels as the
# inspection scale so the existing "School rating" filter UI needs no
# changes. Thresholds are illustrative, chosen from the actual national
# distribution (roughly even-ish population split), not an official scale --
# Scottish Government does not publish a school-rating band for this metric.
ATTAINMENT_BANDS = [
    (60, "Excellent"),
    (45, "Very Good"),
    (35, "Good"),
    (25, "Satisfactory"),
    (15, "Weak"),
    (0, "Unsatisfactory"),
]


def band_for_percent(pct):
    for threshold, label in ATTAINMENT_BANDS:
        if pct >= threshold:
            return label
    return ATTAINMENT_BANDS[-1][1]


def load_locations():
    features = []
    for fname in ("schools_scotland_page1.geojson", "schools_scotland_page2.geojson"):
        with open(os.path.join(RAW, fname)) as f:
            features.extend(json.load(f)["features"])

    schools = {}
    skipped_no_sector = 0
    for feat in features:
        p = feat["properties"]
        sector = SECTOR_MAP.get(p["SchoolType"])
        if sector is None:
            # "Special" schools and any other non primary/secondary type are
            # out of scope per the plan.
            skipped_no_sector += 1
            continue
        lat, lng = p.get("Latitude"), p.get("Longitude")
        if lat is None or lng is None:
            continue
        seed = int(p["SeedCode"])
        address_parts = [
            p.get("AddressLin"),
            p.get("AddressL_1"),
            p.get("AddressL_2"),
        ]
        address = ", ".join(
            part for part in address_parts if part and part.strip() and part.strip() != "0"
        )
        if p.get("PostCode"):
            address = f"{address}, {p['PostCode']}" if address else p["PostCode"]

        website = p.get("WebsiteAdd") or None
        if website and not website.startswith(("http://", "https://")):
            website = f"http://{website}"

        pupil_roll = p.get("Pupil_Roll")
        fte_teachers = p.get("FTE_Teache")

        schools[p["SchUID"]] = {
            "id": p["SchUID"],
            "seedCode": seed,
            "name": p["SchoolName"],
            "sector": sector,
            "denomination": p.get("Denominati") or "Not specified",
            "localAuthority": p["LAName"],
            "address": address,
            "lat": lat,
            "lng": lng,
            "email": p.get("Email") or None,
            "phone": p.get("PhoneNumbe") or None,
            "website": website,
            "pupilRoll": int(pupil_roll) if pupil_roll not in (None, "", 0) else None,
            "fteTeachers": round(fte_teachers, 1) if fte_teachers not in (None, "", 0) else None,
            "rating": {"hasData": False},
        }
    return schools


def slugify(text):
    slug = text.lower().strip()
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    return slug.strip("-")


def assign_page_urls(schools):
    """Sets school['pageUrl'] = 'schools/<la-slug>/<school-slug>.html', appending
    the SEED code to the school slug only when two schools in the same local
    authority would otherwise collide (e.g. same name on split campuses)."""
    la_school_slugs = {}
    for school in schools.values():
        la_slug = slugify(school["localAuthority"])
        school_slug = slugify(school["name"])
        la_school_slugs.setdefault((la_slug, school_slug), []).append(school)

    for (la_slug, school_slug), group in la_school_slugs.items():
        for school in group:
            # SEED code alone doesn't disambiguate: a primary and secondary
            # sharing one campus can share a SEED code (only the SchUID's
            # P/S/SP suffix differs), so use the full id, not seedCode.
            final_slug = (
                school_slug if len(group) == 1 else f"{school_slug}-{school['id'].lower()}"
            )
            school["pageUrl"] = f"schools/{la_slug}/{final_slug}.html"


def load_ratings():
    """Loads inspection QI grades from the 2025 FOI release (all inspections
    up to 1 April 2025). Keeps only the most recent inspection per school.
    QI columns with value 0 mean "not graded under this framework" and are
    excluded from scoring."""
    wb = openpyxl.load_workbook(
        os.path.join(RAW, "foi_inspections_2025.xlsx"), data_only=True
    )
    ws = wb["Public and Grant aided schools"]
    rows = list(ws.iter_rows(values_only=True))

    # Row 0 is a title, row 1 is headers. QI columns are indices 5–13.
    headers = rows[1]
    qi_columns = [
        (idx, str(name).strip())
        for idx, name in enumerate(headers)
        if name and str(name).strip().startswith("QI")
    ]

    latest_by_seed = {}
    for row in rows[2:]:
        seed_raw = row[0]
        if not isinstance(seed_raw, (int, float)):
            continue
        seed = int(seed_raw)
        inspection_date_raw = row[3]
        # Skip rows with no valid inspection date (e.g. "not since opened ...")
        if not inspection_date_raw or not hasattr(inspection_date_raw, "date"):
            if not isinstance(inspection_date_raw, str) or not inspection_date_raw[:4].isdigit():
                continue

        qi_scores = {}
        for idx, qi_name in qi_columns:
            val = row[idx]
            if isinstance(val, (int, float)) and int(val) != 0:
                qi_number = qi_name.replace("QI", "").strip()
                qi_scores[qi_number] = int(val)

        if not qi_scores:
            continue

        existing = latest_by_seed.get(seed)
        if existing is None or inspection_date_raw > existing["inspectionDate"]:
            latest_by_seed[seed] = {
                "seedCode": seed,
                "inspectionDate": inspection_date_raw,
                "qiScores": qi_scores,
            }

    return latest_by_seed


def load_school_stats_2025():
    """Loads fresh pupil roll and teacher FTE from School Level Summary
    Statistics 2025. Returns dict keyed by SEED code (int)."""
    wb = openpyxl.load_workbook(
        os.path.join(RAW, "school_level_stats_2025.xlsx"), data_only=True
    )
    ws = wb["2025 School Level Statistics"]
    rows = list(ws.iter_rows(values_only=True))
    # Row 0 is a title, row 1 is headers.
    by_seed = {}
    for row in rows[2:]:
        seed_raw = row[1]
        if not seed_raw:
            continue
        try:
            seed = int(seed_raw)
        except (ValueError, TypeError):
            continue
        fte = row[4]
        roll = row[5]
        by_seed[seed] = {
            "pupilRoll": int(roll) if isinstance(roll, (int, float)) and roll else None,
            "fteTeachers": round(float(fte), 1) if isinstance(fte, (int, float)) and fte else None,
        }
    return by_seed


def apply_school_stats(schools, stats_by_seed):
    """Overwrites pupilRoll and fteTeachers with 2025 values where available."""
    matched = 0
    for school in schools.values():
        rec = stats_by_seed.get(school["seedCode"])
        if rec is None:
            continue
        if rec["pupilRoll"] is not None:
            school["pupilRoll"] = rec["pupilRoll"]
        if rec["fteTeachers"] is not None:
            school["fteTeachers"] = rec["fteTeachers"]
        matched += 1
    return matched


def apply_ratings(schools, ratings_by_seed):
    matched = 0
    for school in schools.values():
        rec = ratings_by_seed.get(school["seedCode"])
        if rec is None:
            continue
        qi_scores = rec["qiScores"]
        average = sum(qi_scores.values()) / len(qi_scores)
        score = (average - 1) / (6 - 1)
        label = GRADE_LABELS[round(average)]
        inspection_date = rec["inspectionDate"]
        if hasattr(inspection_date, "date"):
            date_str = inspection_date.date().isoformat()
        elif isinstance(inspection_date, str):
            date_str = inspection_date[:10]
        else:
            date_str = str(inspection_date)
        school["rating"] = {
            "hasData": True,
            "metric": "inspection",
            "inspectionDate": date_str,
            "qiScores": qi_scores,
            "averageScore": round(average, 2),
            "score": round(score, 3),
            "label": label,
        }
        matched += 1
    return matched


def load_attainment():
    """Seed code -> {percent, year}, preferring 2024-25 over 2023-24 where
    both exist. Suppressed cells ('*' small-cohort, '#' not applicable) are
    skipped, same as a school having no row at all."""
    sources = [
        ("2024-25", "breadth_and_depth_2024-25.json"),
        ("2023-24", "breadth_and_depth_2023-24.json"),
    ]
    by_seed = {}
    for year, fname in sources:
        with open(os.path.join(RAW, fname)) as f:
            bindings = json.load(f)["results"]["bindings"]
        for row in bindings:
            pct_raw = row["percent"]["value"]
            if not pct_raw.isdigit():
                continue
            seed = int(row["seedCode"]["value"])
            if seed not in by_seed:
                by_seed[seed] = {"percent": int(pct_raw), "year": year}
    return by_seed


def apply_attainment(schools, attainment_by_seed):
    matched = 0
    for school in schools.values():
        if school["sector"] != "secondary":
            continue
        rec = attainment_by_seed.get(school["seedCode"])
        if rec is None:
            continue
        pct = rec["percent"]
        school["rating"] = {
            "hasData": True,
            "metric": "attainment",
            "year": rec["year"],
            "percent": pct,
            "score": round(pct / 100, 3),
            "label": band_for_percent(pct),
        }
        matched += 1
    return matched


def main():
    schools = load_locations()

    stats_2025 = load_school_stats_2025()
    stats_matched = apply_school_stats(schools, stats_2025)
    print(f"Updated pupil roll / teacher FTE from 2025 stats for {stats_matched} schools")

    ratings_by_seed = load_ratings()
    inspection_matched = apply_ratings(schools, ratings_by_seed)

    attainment_by_seed = load_attainment()
    attainment_matched = apply_attainment(schools, attainment_by_seed)

    assign_page_urls(schools)

    out = sorted(schools.values(), key=lambda s: (s["localAuthority"], s["name"]))

    with open(OUT_PATH, "w") as f:
        json.dump(out, f, indent=2)

    by_sector = {}
    rating_metric_counts = {}
    for s in out:
        by_sector[s["sector"]] = by_sector.get(s["sector"], 0) + 1
        if s["rating"]["hasData"]:
            metric = s["rating"]["metric"]
            rating_metric_counts[metric] = rating_metric_counts.get(metric, 0) + 1

    total_with_rating = sum(rating_metric_counts.values())
    print(f"Wrote {len(out)} schools to {OUT_PATH}")
    print(f"  by sector: {by_sector}")
    print(f"  with rating data: {total_with_rating} ({total_with_rating / len(out):.0%})")
    print(f"    inspection-based: {rating_metric_counts.get('inspection', 0)}")
    print(f"    attainment-based (secondary): {attainment_matched}")
    print(f"  generated: {date.today().isoformat()}")


if __name__ == "__main__":
    main()
